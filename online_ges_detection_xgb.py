# coding = utf8

from collections import deque
import datetime

import numpy as np
import xgboost as xgb
import matplotlib.pyplot as plt

from pandas import DataFrame
from pandas.core.common import flatten
from scipy import signal
from openpyxl import load_workbook


def band_filter(data):
    [b, a] = signal.butter(3, 0.00001, btype='highpass')
    filtered_data = signal.filtfilt(b, a, data, padlen=0)
    [b, a] = signal.butter(3, 0.3, btype='lowpass')
    filtered_data = signal.filtfilt(b, a, filtered_data, padlen=0)
    return np.abs(filtered_data)


def rebuild_channel_data(data):
    data = np.array(data)
    return DataFrame(data, columns=['ch%d' % (i + 1) for i in range(len(data[0]))])


def get_moving_avg(data):
    """
    Input:
        data (DataFrame): each column corresponds a channel, with each row represents a sampled point
    Output:
        data (list): a list of the rolling average of the maximum data at each sampled point
    """
    data = data.max(axis=1)
    data = list(flatten((data.rolling(2).mean()[2:len(data)]).values))
    return data


def segmentation(data, threshold, window_length):
    """
    Input:
        data (list): dim = num_of_samples x 1
        threshold (int)
        window_length (int)
    Output:
        new_start_list (list): dim = num_of_cuts x 1
        new_end_list (list): dim = num_of_cuts x 1
    """
    raw_start_list = len(data) * [-1]
    end_list = []
    start_list = []
    period_list = []
    for i in range(len(data)):
        if data[i] > threshold:
            raw_start_list[i] = i
    for m in range(len(data)):
        check = raw_start_list[m]
        if check > 0:
            add = 0
            for n in range(m + 1, len(raw_start_list)):
                if raw_start_list[n] > 0:
                    if raw_start_list[n] - check < window_length:
                        check = raw_start_list[n]
                        raw_start_list[n] = -999
                        add = 0
                    else:
                        if check > 0 and raw_start_list[n] > 0:
                            end_list.append(check + window_length)
                            break
                else:
                    add += 1
                    if add >= window_length:
                        end_list.append(check)
                        break
    for item in raw_start_list:
        if item > 0:
            start_list.append(item)

    new_start_list = []
    new_end_list = []
    for i in range(len(end_list) - 1):
        period_list.append(end_list[i] - start_list[i])
    for i in range(len(period_list)):
        if period_list[i] > 300:
            new_start_list.append(start_list[i])
            new_end_list.append(end_list[i])
    return new_start_list, new_end_list


class OnlineGestureDetection(object):
    def __init__(self, model_path, param):
        self.model = xgb.Booster(model_file=model_path)
        self.buffer = deque(maxlen=param['buffer_length'])
        self.threshold = param['threshold']
        self.window_length = param['window_length']
        self.previous_prediction = None

    def take_in_data(self, data):
        self.buffer.append(data)

    def predict(self):

        data = np.array(self.buffer)
        plt.plot(data)
        plt.ylim([0, 200])
        plt.show()
        filtered_data = np.zeros([len(data), len(data[0])])
        for i in range(len(data[0])):
            filtered_data[:, i] = np.array(band_filter(np.array(data)[:, i]))

        start_list, end_list = \
            segmentation(get_moving_avg(rebuild_channel_data(filtered_data)), self.threshold, self.window_length)
        if start_list and end_list:
            start = start_list[-1]
            end = end_list[-1]
            filtered_data = filtered_data[start:end]
            mav = np.expand_dims(np.mean(filtered_data, axis=0), axis=0)
            predict = self.model.predict(xgb.DMatrix(mav))[0]
            self.previous_prediction = predict
        else:
            predict = self.previous_prediction
        return predict


if __name__ == '__main__':
    param = {'threshold': 2,
             'window_length': 20,
             'buffer_length': 2000}
    detect = OnlineGestureDetection('xgb.model', param)
    name = 'train/training3.xlsx'
    workbook = load_workbook(name).active
    cnt = 0
    for row in workbook.iter_rows(min_row=2, max_col=8, max_row=workbook.max_row, values_only=True):
        t_start = datetime.datetime.now()

        detect.take_in_data(list(row))
        if cnt % 200 == 0:
            output = detect.predict()
        t_end = datetime.datetime.now()
        print("sample: ", cnt, "  classification: ", output, "  time bin: ", (t_end - t_start).total_seconds(), 's')
        cnt += 1
