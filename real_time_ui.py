# coding = utf8

from ctypes import *
from collections import deque
import sys
import datetime
import random

import openpyxl
from pyqtgraph.Qt import QtCore, QtGui
import pyqtgraph as pg
import numpy as np
import matplotlib.pyplot as plt

from PyQt5.QtCore import Qt, QThread, pyqtSignal
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QLabel, QVBoxLayout, QHBoxLayout
from pyqtgraph import PlotWidget, PlotItem
from openpyxl import load_workbook

from online_ges_detection_dnn import OnlineGestureDetection


class sub_Struct(Structure):
    _fields_ = [
        ("sub_test_int", c_int),
        ("sub_test_char_arr", c_char * 300)
    ]


class NET_DVR_ALARMER(Structure):
    _fields_ = [
        ("test_int", c_int),
        ("char_array", c_char * 20000),
        ("test_sub_struct", sub_Struct),
        ("byte_test_p", POINTER(c_byte))
    ]


CALLFUNC = CFUNCTYPE(c_void_p, POINTER(NET_DVR_ALARMER))


class ComThread(QThread):
    expSignal = pyqtSignal(list)

    def __init__(self, parent=None):
        super(ComThread, self).__init__(parent)

    def callback_msg(self, type_struct):
        str_tem = str(type_struct.contents.char_array.decode())
        data_tem = []
        str_list_tem = str_tem.split(',')
        str_list_tem.pop()
        int_tem_array = []
        for index, item in enumerate(str_list_tem):
            item = int(item)
            int_tem_array.append(item)
            if (index + 1) % 8 == 0:
                data_tem.append(int_tem_array)
                int_tem_array = []
        self.expSignal.emit(data_tem)

    def run(self):
        my_lib = windll.LoadLibrary('FetchGForceData64.dll')
        my_lib.check(CALLFUNC(self.callback_msg), None)


class Demo(QWidget):

    def __init__(self, param):
        super(Demo, self).__init__()
        self.detect = OnlineGestureDetection('dnn.pt', param)
        self.if_predict = param['if_predict']
        self.resize(3200, 1800)

        self.save_button = QPushButton(self)
        self.save_button.setText('Save Data')
        self.save_button.move(2410, 1210)

        self.save_button.clicked.connect(self.save_data)

        self.p1 = PlotWidget(self)
        self.p1.setGeometry(QtCore.QRect(10, 10, 780, 580))
        self.p2 = PlotWidget(self)
        self.p2.setGeometry(QtCore.QRect(810, 10, 780, 580))
        self.p3 = PlotWidget(self)
        self.p3.setGeometry(QtCore.QRect(1610, 10, 780, 580))
        self.p4 = PlotWidget(self)
        self.p4.setGeometry(QtCore.QRect(2410, 10, 780, 580))
        self.p5 = PlotWidget(self)
        self.p5.setGeometry(QtCore.QRect(10, 610, 780, 580))
        self.p6 = PlotWidget(self)
        self.p6.setGeometry(QtCore.QRect(810, 610, 780, 580))
        self.p7 = PlotWidget(self)
        self.p7.setGeometry(QtCore.QRect(1610, 610, 780, 580))
        self.p8 = PlotWidget(self)
        self.p8.setGeometry(QtCore.QRect(2410, 610, 780, 580))
        self.p9 = PlotWidget(self)
        self.p9.setGeometry(QtCore.QRect(10, 1210, 2380, 580))
        self.p1.setXRange(0, 100)
        self.p2.setXRange(0, 100)
        self.p3.setXRange(0, 100)
        self.p4.setXRange(0, 100)
        self.p5.setXRange(0, 100)
        self.p6.setXRange(0, 100)
        self.p7.setXRange(0, 100)
        self.p8.setXRange(0, 100)
        self.p1.setYRange(80, 160)
        self.p2.setYRange(80, 160)
        self.p3.setYRange(80, 160)
        self.p4.setYRange(80, 160)
        self.p5.setYRange(80, 160)
        self.p6.setYRange(80, 160)
        self.p7.setYRange(80, 160)
        self.p8.setYRange(80, 160)
        self.p9.setYRange(0, 1)
        self.curve1 = self.p1.plot(np.random.normal(size=100))
        self.curve2 = self.p2.plot(np.random.normal(size=100))
        self.curve3 = self.p3.plot(np.random.normal(size=100))
        self.curve4 = self.p4.plot(np.random.normal(size=100))
        self.curve5 = self.p5.plot(np.random.normal(size=100))
        self.curve6 = self.p6.plot(np.random.normal(size=100))
        self.curve7 = self.p7.plot(np.random.normal(size=100))
        self.curve8 = self.p8.plot(np.random.normal(size=100))
        self.bar = pg.BarGraphItem(x=range(1, 4), height=[[1/3]] * 3, width=0.6, brush='w')
        self.p9.addItem(self.bar)

        self.my_thread = ComThread()
        self.my_thread.expSignal.connect(self.receive_data)
        self.my_thread.start()

    def receive_data(self, data_tem):
        self.detect.take_in_multiple_data(data_tem)

        self.curve1.setData(np.array(self.detect.buffer)[-100:, 0])
        self.curve2.setData(np.array(self.detect.buffer)[-100:, 0])
        self.curve3.setData(np.array(self.detect.buffer)[-100:, 0])
        self.curve4.setData(np.array(self.detect.buffer)[-100:, 0])
        self.curve5.setData(np.array(self.detect.buffer)[-100:, 0])
        self.curve6.setData(np.array(self.detect.buffer)[-100:, 0])
        self.curve7.setData(np.array(self.detect.buffer)[-100:, 0])
        self.curve8.setData(np.array(self.detect.buffer)[-100:, 0])
        if self.if_predict:
            predict, softmax_scores = self.detect.predict()
            self.bar.setOpts(height=np.array(self.bar.getData()[1]).flatten() * 0.9 + softmax_scores * 0.1)



    def save_data(self):
        data = list(self.detect.buffer)
        print(len(data))
        wb = openpyxl.Workbook()
        ws = wb.active

        for item in data:
            ws.append(item)
        wb.save('train/new.xlsx')

if __name__ == '__main__':
    real_time = True
    if real_time:
        param = {'threshold': 2,
                 'window_length': 20,
                 'buffer_length': 200,
                 'effective_zone': 100,
                 'filter': True,
                 'if_predict': True}
        app = QApplication(sys.argv)
        demo = Demo(param)
        demo.show()
        sys.exit(app.exec())
    else:
        param = {'threshold': 2,
                 'window_length': 20,
                 'buffer_length': 200,
                 'effective_zone': 100,
                 'filter': True,
                 'if_predict': True}
        detect = OnlineGestureDetection('dnn.pt', param)

        name = 'train/test2.xlsx'
        workbook = load_workbook(name).active
        cnt = 0
        lines = [[]] * 8
        plt.show()
        fig = plt.figure(constrained_layout=False, figsize=(11, 6))
        grid_spec = fig.add_gridspec(nrows=3, ncols=4, left=0.05, bottom=0.05, right=0.95, top=0.95)

        for row in range(2):
            for col in range(4):
                ax = fig.add_subplot(grid_spec[row, col])
                ax.set_title('Channel %d' % (row * 4 + col + 1))
                ax.set_xlim(0, 100)
                ax.set_ylim(80, 160)
                lines[row * 4 + col], = ax.plot([], [], color=(
                    random.uniform(0, 1), random.uniform(0, 1), random.uniform(0, 1)))
        ax = fig.add_subplot(grid_spec[2, 0:3])
        bar = ax.bar(range(1, 9), [0] * 8)
        ax.set_title('Gesture Scores')
        ax.set_ylim([0, 1])

        ax = fig.add_subplot(grid_spec[2, 3])
        ax.set_title('Classification')

        for row in workbook.iter_rows(min_row=2, max_col=8, max_row=workbook.max_row, values_only=True):
            t_start = datetime.datetime.now()
            detect.take_in_data((list(row)))
            predict, softmax_scores = detect.predict()
            t_end = datetime.datetime.now()

            for idx in range(8):
                lines[idx].set_xdata(range(len(np.array(detect.buffer)[-100:, idx])))
                lines[idx].set_ydata(np.array(detect.buffer)[-100:, idx])
            for i in range(8):
                bar[i].set_height(0.1 * softmax_scores[i] + 0.9 * bar[i].get_height())
            plt.draw()
            plt.pause(1e-17)

            print("sample: ", cnt, "  classification: ", predict, "  time bin: ", (t_end - t_start).total_seconds(),
                  's')
            cnt += 1
