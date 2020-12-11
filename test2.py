# coding=utf8
from PyQt5.QtCore import *
from PyQt5 import QtGui, uic

import threading, sys, math, os

import pyqtgraph as pg
import numpy as np
from PyQt5.QtGui import QWidget, QApplication, QFrame, QGridLayout, QVBoxLayout, QLabel, QPushButton
from PyQt5.QtCore import Qt, QTimerpy
# from twisted.internet import reactor, defer, threads, error, task

# 读取ini
import configparser, codecs

import serial
import time

# 读取excel
from openpyxl import Workbook, load_workbook

# 自定义模块
from train import *

# UDP发包
from package import *
import socket

# 带通滤波
from scipy import signal
from scipy import sparse
from scipy.fftpack import rfft, irfft, fftfreq

from scipy.sparse.linalg import spsolve

"""Python 3 不需要check
#check
def QString2PyString(qStr):
    # # QString，如果内容是中文，则直接使用会有问题，要转换成 python string
    return str(qStr.toUtf8(), 'utf-8', 'ignore')
"""

# 注册回调函数
from ctypes import *


# C++回调函数中的子结构体
class sub_Struct(Structure):
    _fields_ = [
        ("sub_test_int", c_int),
        ("sub_test_char_arr", c_char * 300)
    ]


# 回调函数结构体
class NET_DVR_ALARMER(Structure):
    _fields_ = [
        ("test_int", c_int),
        ("char_array", c_char * 20000),
        ("test_sub_struct", sub_Struct),
        ("byte_test_p", POINTER(c_byte))
    ]


# 回调函数的参数
CALLFUNC = CFUNCTYPE(c_void_p, POINTER(NET_DVR_ALARMER))


class comThread(QThread):
    expSignal = pyqtSignal(list)
    def __init__(self, parent=None):
        super(comThread, self).__init__(parent)
        self.startRecord = False
        self.stopRecord = False
        self.port = ''
        self.baudrate = 0
        self.l_serial = None
        self.ID = None
        self.data = []

    def SendDate(self, i_msg, send):
        lmsg = ''
        isOK = False
        if isinstance(i_msg):
            lmsg = i_msg.encode('gb18030')
        else:
            lmsg = i_msg
        try:
            # 发送数据到相应的处理组件
            self.l_serial.write(send)
        except Exception as ex:
            pass;
        return isOK

    def callbackMsg(self, type_struct):

        # dataTem = []
        strTem = [str(type_struct.contents.char_array.decode())]
        # print(strTem)
        """
        strlistTem = strTem.split(',')

        strlistTem.pop()
        intTemArray = []
        for index,item in enumerate(strlistTem):
            item = int(item)
            intTemArray.append(item)
            if (index + 1) % 8 == 0:
                dataTem.append(intTemArray)
                intTemArray = []

        #print(dataTem)
        self.expSignal.emit(dataTem)
        """
        self.expSignal.emit(strTem)
        # return str(type_struct.contents.char_array.decode())
        # print("结构体读取，char*变量：" + str(type_struct.contents.char_array.decode()))

    def run(self):
        try:
            self.startFun()
            pass
        except Exception as ex:
            print((str(ex)))
        self.secondReader()

    def startFun(self):
        mylib = windll.LoadLibrary("FetchGForceData.dll")
        mylib.check(CALLFUNC(self.callbackMsg), None)


'''加载对话框资源'''
Ui_list = []
for i in range(1):
    _name = 'UI/ui%d.ui' % i
    Ui_list.append(uic.loadUiType(_name))


class MyApp(QtGui.QMainWindow, Ui_list[0][0]):
    msgSignal = pyqtSignal(str)
    ptcSignal = pyqtSignal(int)

    def __init__(self):
        QtGui.QMainWindow.__init__(self)
        Ui_list[0][0].__init__(self)
        self.setupUi(self)

        self.openBtn.clicked.connect(self.openCOM)
        self.trainBtn.clicked.connect(self.pre_startTrainFun)
        self.ensureBtn.clicked.connect(self.ensureFun)
        self.testBtn.clicked.connect(self.testBeginFun)

        self.msgSignal.connect(self.showMsg)
        self.ptcSignal.connect(self.changePic)

        # params
        self.comOpened = False  # 判断端口是否已经打开
        self.myQWidgetCreated = False  # 判断是否首次创建QWidget窗口
        self.msgIndex = 1  # msg序号
        self.threashold = ''  # 解算出的基线阈值
        self.startbaseline = False  # 是否开始基线记录
        self.allowDraw = True  # 是否绘图
        self.baselineArray = [0, 0, 0, 0, 0, 0, 0, 0]  # 用来计算baseline的数组
        self.gestureSampleDone = True  # 是否已经完成一组训练数据采集
        self.gestureSampleIndex = -1  # 用来标记是哪种类型的训练数据
        self.trainingData = [[], [], [], [], [], [], [], []]  # 存放训练数据
        self.model = None  # 存放训练模型
        self.isTesting = False  # 是否进行测试
        self.bufferArray = pd.DataFrame(columns=['ch1', 'ch2', 'ch3', 'ch4'])  # 用来存放实时动作检测的缓冲区
        self.startDetecting = False  # 是否正在进行一个动作检测
        self.resultIndex = -1  # 检测的结果
        self.actinLatency = 0  # 执行动作的潜伏期
        self.dataListBuffer = []  # 用来缓冲实时动作检测的原始数据
        self.trained = False  # 此次使用是否重新进行了训练

        self.allowBaselineRecording = True
        self.allowFiltering = False
        self.faking = False  # True
        self.data_mean_semg = []

        self.intiFun()

    def baseline_als(self, y, lam, p,
                     niter=10):  # baseline_als(y, 10000, 0.05)  0.001 ≤ p ≤ 0.1 is a good choice (for a signal with positive peaks) and 10^2 ≤ λ ≤ 10^9
        L = len(y)
        D = sparse.csc_matrix(np.diff(np.eye(L), 2))
        w = np.ones(L)
        z = None
        for i in range(niter):
            W = sparse.spdiags(w, 0, L, L)
            Z = W + lam * D.dot(D.transpose())
            z = spsolve(Z, w * y)
            w = p * (y > z) + (1 - p) * (y < z)
        return z

    def getcode(self):
        userconfig = configparser.ConfigParser()
        userconfig.readfp(open('./config.ini'))
        self.threashold = userconfig.get('modelInfo', 'threashold')
        # print userconfig.get('modelInfo', 'threashold')
        # print self.threashold

    def setCode(self, _section, _option, _value):
        userconfig = configparser.ConfigParser()
        userconfig.read('./config.ini')
        userconfig.set(_section, _option, _value)
        fp = codecs.open('./config.ini', 'wb', 'utf-8')
        userconfig.write(fp)

    def intiFun(self):
        pix0 = QtGui.QPixmap('UI\images\g0.png')
        self.gesTurelabel.setPixmap(pix0)

        # 绘图起始设置
        self.myQWidget = QWidget(self)
        self.myQWidget.setGeometry(92, 22, 1037, 477)
        self.myQWidget.setWindowTitle("Online EMG ploting")
        self.myQWidget.gridLayout = QGridLayout(self.myQWidget)
        self.myQWidget.frame1 = QFrame(self)
        self.myQWidget.frame1.setFrameShape(QFrame.Panel)
        self.myQWidget.frame1.setFrameShadow(QFrame.Plain)
        self.myQWidget.frame1.setLineWidth(2)
        self.myQWidget.frame1.setStyleSheet("background-color:rgb(0,255,255);")

        self.myQWidget.frame2 = QFrame(self)
        self.myQWidget.frame2.setFrameShape(QFrame.Panel)
        self.myQWidget.frame2.setFrameShadow(QFrame.Plain)
        self.myQWidget.frame2.setLineWidth(2)
        self.myQWidget.frame2.setStyleSheet("background-color:rgb(0,255,255);")

        self.myQWidget.frame3 = QFrame(self)
        self.myQWidget.frame3.setFrameShape(QFrame.Panel)
        self.myQWidget.frame3.setFrameShadow(QFrame.Plain)
        self.myQWidget.frame3.setLineWidth(2)
        self.myQWidget.frame3.setStyleSheet("background-color:rgb(0,255,255);")

        self.myQWidget.frame4 = QFrame(self)
        self.myQWidget.frame4.setFrameShape(QFrame.Panel)
        self.myQWidget.frame4.setFrameShadow(QFrame.Plain)
        self.myQWidget.frame4.setLineWidth(2)
        self.myQWidget.frame4.setStyleSheet("background-color:rgb(0,255,255);")

        self.myQWidget.gridLayout.addWidget(self.myQWidget.frame1, 0, 0, 1, 1)
        self.myQWidget.gridLayout.addWidget(self.myQWidget.frame2, 0, 1, 1, 1)
        self.myQWidget.gridLayout.addWidget(self.myQWidget.frame3, 1, 0, 1, 1)
        self.myQWidget.gridLayout.addWidget(self.myQWidget.frame4, 1, 1, 1, 1)
        self.myQWidget.setLayout(self.myQWidget.gridLayout)

        self.array1 = []
        self.array2 = []
        self.array3 = []
        self.array4 = []
        self.t = []

        self.array1Old = []
        self.array2Old = []
        self.array3Old = []
        self.array4Old = []

        # 读取配置表
        t2 = threading.Thread(target=self.getcode, args=())
        t2.start()
        while True:
            if not t2.is_alive():
                break
        temArray = self.threashold.split(',')
        for index, item in enumerate(temArray):
            if item:
                self.baselineArray[index] = int(item)
        # print(self.baselineArray)

        self.UDPserverInit()

    def testBeginFun(self):
        if not self.trained:
            try:
                import xgboost as xgb
                self.model = xgb.Booster(model_file='xgb.model')
                msgTem = '成功加载训练模型'
                self.msgSignal.emit(msgTem)
            except:
                msgTem = '未能加载已存在的模型,请重新进行训练'
                self.msgSignal.emit(msgTem)

        self.isTesting = True
        # t3 = threading.Thread(target=self.actionSummit, args=())
        # t3.start()

    def ensureFun(self):
        self.gestureSampleDone = True

    def changePic(self, index):
        strTem = 'UI\images\g%d.png' % index
        mapTem = QtGui.QPixmap(strTem)
        self.gesTurelabel.setPixmap(mapTem)

    def showMsg(self, msg):
        msg = '%d. %s' % (self.msgIndex, msg)
        self.msgIndex += 1
        self.myMsgBox.addItem(msg)

    # 准备开始训练
    def pre_startTrainFun(self):
        self.trainBtn.setEnabled(False)
        t1 = threading.Thread(target=self.startTrainFun, args=())
        t1.start()

    def lantencyWork(self, lantency):
        tCheck = int(round(time.time() * 1000))
        while 1:
            tNow = int(round(time.time() * 1000))
            if tNow >= tCheck + lantency:
                return

    def startTrainFun(self):

        self.testBtn.setEnabled(False)

        msgTem = '准备开始训练'
        self.msgSignal.emit(msgTem)

        if self.allowBaselineRecording:
            self.ensureBtn.setEnabled(False)
            # 记录基线，并通过基线计算阈值
            self.lantencyWork(1000)
            msgTem = '开始基线采集，约10s'
            self.msgSignal.emit(msgTem)
            msgTem = '请放松被测肢体，保持静止'
            self.msgSignal.emit(msgTem)
            self.startbaseline = True
            # self.lantencyWork(10000)
            self.startbaseline = False

            # maxNum = max(self.baselineArray)
            # sumNum = sum(self.baselineArray)
            # sumNum = sumNum - maxNum
            temArray = [str(item) for item in self.baselineArray]
            self.threashold = ','.join(temArray)

            # self.threashold = self.fakebaseline()

            msgTem = '基线采集结束，基线阈值为 %s' % (self.threashold)
            self.msgSignal.emit(msgTem)

            self.setCode('modelInfo', 'threashold', str(self.threashold))

            self.ensureBtn.setEnabled(True)

        # textArray = [u'握拳',u'张手',u'外翻',u'内弯']
        textArray = ['握拳', '张手', '外翻', '内弯', '捏手', '手', '比六', 'OK']
        self.trainingData = [[], [], [], [], [], [], [], []]
        for i in range(8):
            textTem = textArray[i]
            self.lantencyWork(1000)
            msgTem = '开始采集第%d个动作:%s' % (i + 1, textTem)
            self.msgSignal.emit(msgTem)
            self.lantencyWork(100)
            self.ptcSignal.emit(i)
            msgTem = '请按右侧图片提示做出相应手势'
            self.msgSignal.emit(msgTem)
            self.lantencyWork(100)
            if i == 0:
                msgTem = '请采集15-20个动作，采集完毕按右方确认按钮确认'
                self.msgSignal.emit(msgTem)

            self.gestureSampleIndex = i  # 标识出采集的动作编号
            self.gestureSampleDone = False  # 开始采集

            while True:
                if self.gestureSampleDone:
                    msgTem = '%s 训练集已经被确认' % (textTem)
                    self.msgSignal.emit(msgTem)
                    break

            if not self.faking:
                self.restoreTrainingData()

        msgTem = '训练数据已经采集完毕，请等待训练模型生成'
        self.msgSignal.emit(msgTem)

        if self.faking:
            self.trainingData = self.faketraining()
        # print self.trainingData

        '''
        try:
            accuracy, self.model = trainFun(self.trainingData, self.threashold)
            msgTem = u'xgboost模型已经成功生成'
            self.msgSignal.emit(msgTem)
            msgTem = u'预测准确率：%.2f %% '% accuracy
            self.msgSignal.emit(msgTem)
            self.trained = True
        except:
            msgTem = u'xgboost模型生成失败'
            self.msgSignal.emit(msgTem)
        '''

        accuracy, self.model = trainFun(self.trainingData, self.baselineArray, self.allowFiltering)
        msgTem = 'xgboost模型已经成功生成'
        self.msgSignal.emit(msgTem)
        msgTem = '预测准确率：%.2f %% ' % accuracy
        self.msgSignal.emit(msgTem)
        self.trained = True

        self.testBtn.setEnabled(True)

    def fakebaseline(self):
        workbook = load_workbook('baselinefake.xlsx')
        sheet1 = workbook.worksheets[0]
        array = [0, 0, 0, 0]
        for i in range(2, sheet1.max_row):
            for j in range(4):
                valueTem = sheet1.cell(row=i, column=1 + j).value
                array[j] = (array[j] + valueTem) / 2

        maxNum = max(array)
        minNum = min(array)
        sumNum = sum(array)
        sumNum = sumNum - maxNum - minNum
        numNew = (sumNum / (len(array) - 2)) * 2

        return numNew

    def bandfilter(self, data):
        b, a = signal.butter(3, 0.05, 'highpass')
        filtedData = signal.filtfilt(b, a, data, padlen=0)
        return filtedData

    # 用来激发真正的绘图函数
    def plotDataReal(self, dataList):  # check

        for index, item in enumerate(dataList):
            self.array1.append(item[4])
            self.array2.append(item[5])
            self.array3.append(item[6])
            self.array4.append(item[7])
            self.t.append(index + self.myQWidget.pha)

        # if self.myQWidget.pha % 50 == 0:
        # self.array4 = self.baseline_als(self.array4, 10, 0.1)
        self.myQWidget.curve1.setData(self.t, self.array1)
        self.myQWidget.curve2.setData(self.t, self.array2)
        self.myQWidget.curve3.setData(self.t, self.array3)
        self.myQWidget.curve4.setData(self.t, self.array4)

        # self.myQWidget.pha += 10
        self.myQWidget.pha += 20

        if self.myQWidget.pha >= 1000:
            self.myQWidget.pha = 0
            self.array1 = []
            self.array2 = []
            self.array3 = []
            self.array4 = []
            self.t = []

            self.array4Old = []
            self.array3Old = []
            self.array2Old = []
            self.array1Old = []

    def faketraining(self):
        outputlast = []
        if self.allowFiltering:
            for m in range(8):
                name = 'modu%d.xlsx' % m
                workbook = load_workbook(name)
                sheet1 = workbook.worksheets[0]

                outputfilter = [[], [], [], [], [], [], [], []]

                cnt = 0
                for j in range(8):
                    bufferArray = []
                    for i in range(2, sheet1.max_row):
                        valueTem = sheet1.cell(row=i, column=1 + j).value
                        bufferArray.append(valueTem)
                        if cnt >= 10:
                            cnt = 0
                            temlist = self.bandfilter(bufferArray)
                            temArray = temlist.tolist()
                            temArray = [float('%.3f' % item) for item in temArray]
                            outputfilter[j] = outputfilter[j] + temArray

                            bufferArray = []
                        cnt += 1
                outputfilter = self.fourcollistTosimplelist(outputfilter)

                outputlast.append(outputfilter)
        else:
            for i in range(8):
                name = 'modu%d.xlsx' % i
                workbook = load_workbook(name)
                sheet1 = workbook.worksheets[0]

                output = []
                for n in range(2, sheet1.max_row):
                    array = [0, 0, 0, 0, 0, 0, 0, 0]
                    for j in range(8):
                        valueTem = sheet1.cell(row=n, column=1 + j).value
                        array[j] = valueTem
                    output.append(array)
                outputlast.append(output)
        return outputlast

    # 用来存储训练数据
    def restoreTrainingData(self):
        itemNow = self.trainingData[self.gestureSampleIndex]

        workbook = Workbook()
        sheet1 = workbook.worksheets[0]
        # 先在第一行写入表头
        for i in range(8):
            sheet1.cell(row=1, column=1 + i).value = 'ch%d' % (i + 1)
        self.workbook = workbook

        for i, item in enumerate(itemNow):
            for j, num in enumerate(item):
                sheet1.cell(row=i + 2, column=1 + j).value = num

        fileName = 'training%d.xlsx' % self.gestureSampleIndex
        workbook.save(fileName)

    # 调用串口，测试串口
    def comMain(self):
        # port = QString2PyString(self.comInputbox.text())
        port = self.comInputbox.text()

        # 监听端口
        self.comThread = comThread()
        # self.comThread.port = port               #check
        self.comThread.port = 'COM3'  # check
        self.comThread.baudrate = 115200  # check
        self.comThread.sendport = '**1*80*'

        self.comThread.expSignal.connect(self.recieveData)
        self.comThread.start()

    # 结果会转换为X4的二元序列,而输入为简单的4元数序列
    def preFiltering(self, datalist):
        array1 = []
        array2 = []
        array3 = []
        array4 = []
        array5 = []
        array6 = []
        array7 = []
        array8 = []

        arrayTem_1 = []
        arrayTem_2 = []
        arrayTem_3 = []
        arrayTem_4 = []
        arrayTem_5 = []
        arrayTem_6 = []
        arrayTem_7 = []
        arrayTem_8 = []

        for index, item in enumerate(datalist):
            array1.append(item[0])
            array2.append(item[1])
            array3.append(item[2])
            array4.append(item[3])
            array5.append(item[3])
            array6.append(item[3])
            array7.append(item[3])
            array8.append(item[3])

        for i in range(1, 9):
            # exec ('array%d = self.baseline_als(array%d, 10000, 0.001)' % (i, i))
            exec('array%d = self.bandfilter(array%d)' % (i, i))
            exec('arrayTem_%d = arrayTem_%d + array%d.tolist()' % (i, i, i))

        outPut = [arrayTem_1, arrayTem_2, arrayTem_3, arrayTem_4, arrayTem_5, arrayTem_6, arrayTem_7, arrayTem_8]

        for i in range(8):
            outPut[i] = [float('%.3f' % item) for item in outPut[i]]
        return outPut

    # 4列二元数组转4元数数组
    def fourcollistTosimplelist(self, list4):
        lenTem = min(len(list4[0]), len(list4[1]), len(list4[2]), len(list4[3]), len(list4[4]), len(list4[5]),
                     len(list4[6]), len(list4[7]))
        outPut = []
        for i in range(lenTem):
            outPutTem = [[], [], [], [], [], [], [], []]
            for j in range(8):
                item = list4[j][i]
                outPutTem[j] = item
            outPut.append(outPutTem)
        return outPut

    # 4元数数组转4列二元数组
    def simplelistTofourcollist(self, lists):
        outPut = [[], [], [], [], [], [], [], []]
        for i in range(len(lists)):
            for j in range(8):
                item = lists[i][j]
                outPut[j].append(item)
        return outPut

    def recieveData(self, dataList):

        # 是否滤波
        if self.allowFiltering:
            dataList = self.preFiltering(dataList)

        # print dataList

        # 是否绘图
        if self.allowDraw:
            self.plotDataReal(dataList)  # 需要4list

        if self.allowFiltering:
            dataList = self.fourcollistTosimplelist(dataList)  # 先加上这个转换，以后慢慢改

        # 是否解算基线
        if self.startbaseline:
            self.caculateBaseline(dataList)  # 需要list4num
        # 是否记录训练数据
        if not self.gestureSampleDone:
            self.recordingTrainingDataFun(dataList)  # 需要list4num
        # 是否开始测试                                     #需要list4num
        if self.isTesting:
            # 这个datalist必须被接收，否则会造成信号失真
            self.testFunOnline(dataList)

        """
        #print(dataList)
        res_dict = {"Fist": 0, "SpreadFingers": 1, "WaveOut": 2, "WaveIn": 3,
                    "Pinch":4, "Shoot":5, "Relax":6}

        if res_dict[dataList[0]] != 6:
            self.resultIndex = res_dict[dataList[0]]
            self.actionSummit()
        """

    def testFunOnline(self, dataList):
        self.dataListBuffer.extend(dataList)
        # if len(self.dataListBuffer) > 64 and (not self.startDetecting):

        if not self.startDetecting:
            self.startDetecting = True
            dataListTem = self.dataListBuffer[:]
            self.dataListBuffer = []

            t2 = threading.Thread(target=onlineEventDetection,
                                  args=(dataListTem, self.baselineArray, 100, self.model, self))
            t2.start()


    def recordingTrainingDataFun(self, dataList):
        self.trainingData[self.gestureSampleIndex].extend(dataList)

    # 计算baseline，简单地就每个通道计算平均值
    def caculateBaseline(self, dataList):
        for dataTem in dataList:
            for index, item in enumerate(dataTem):
                self.baselineArray[index] = (self.baselineArray[index] + item) / 2

    def openCOM(self):
        if self.comOpened:
            # 如果已经打开，则显示为关闭端口，做关闭端口操作
            self.comOpened = False
            # 停止绘图
            self.timer.stop()

            self.openBtn.setText('Open COM')
        else:
            self.comOpened = True
            # 打开端口
            self.comMain()
            print("已打开串口")
            # print(ID, data)
            # 绘图准备
            self.myQWidget.show()
            self.generate_image()

            self.openBtn.setText('close COM')

    def generate_image(self):
        if self.myQWidgetCreated:
            return
        for i in range(1, 5):
            exec('verticalLayout = QVBoxLayout(self.myQWidget.frame%d)' % i)
            exec('win%d = pg.GraphicsLayoutWidget(self.myQWidget.frame%d)' % (i, i))
            exec('verticalLayout.addWidget(win%d)' % i)
            exec('p%d = win%d.addPlot(title=u"Online EMG Channel %d")' % (i, i, i))
            exec('p%d.showGrid(x=True,y=True)' % i)
            exec("p%d.setLabel(axis='left',text='Amplitude / V')" % i)
            exec("p%d.setLabel(axis='bottom',text='t / s')" % i)
            # exec ('p%d.setYRange(-300,300)'% i)
            if self.allowFiltering:
                exec('p%d.setYRange(-30,30)' % i)
            else:
                exec('p%d.setYRange(100,250)' % i)
            # exec ('p%d.setYRange(-30,30)' % i)
            exec('p%d.setXRange(0, 1000)' % i)
            # exec ('p%d.addLegend()'% i)
            exec("self.myQWidget.curve%d = p%d.plot(pen='r',name='y%d')" % (i, i, i))

        # self.myQWidget.Fs = 200.0 #采样频率
        # self.myQWidget.N = 200    #采样点数
        self.myQWidget.Fs = 400.0  # 采样频率
        self.myQWidget.N = 400  # 采样点数
        # self.myQWidget.f0 = 4.0    #信号频率
        self.myQWidget.f0 = 8.0  # 信号频率
        self.myQWidget.pha = 0  # 初始相位
        self.myQWidget.t = np.arange(self.myQWidget.N) / self.myQWidget.Fs  # 时间向量 1*1024的矩阵

        self.myQWidgetCreated = True



if __name__ == '__main__':
    app = QtGui.QApplication(sys.argv)
    window = MyApp()
    window.show()
    sys.exit(app.exec_())