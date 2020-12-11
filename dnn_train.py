import time
import copy
import os

import numpy as np
from torch.utils.data import TensorDataset
import torch.nn as nn
import torch.optim as optim
import torch
from torch.autograd import Variable
from openpyxl import load_workbook

from dnn_model import DeepNeuralNet

from online_ges_detection_dnn import band_filter


class TrainDNN(object):
    def __init__(self, data_folder_path, param):
        self.data_folder_path = data_folder_path
        self.num_of_ges = param['num_of_ges']
        self.threshold = param['threshold']
        self.window_length = param['window_length']
        self.visualization = param['visualization']
        self.training_iter = param['training_iter']
        self.learning_rate = param['learning_rate']
        self.device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
        self.batch_size = param['batch_size']
        self.array_x = []
        self.array_y = []

    def data_preprocessing(self):
        data_list = []
        data_dict = {}
        x_train = []
        y_train = []

        def n_col_list2simple_list(n_col_list):
            minimum_row = min([len(n_col_list[i]) for i in range(len(n_col_list))])
            simple_list = []
            for i in range(minimum_row):
                temporary_output = len(n_col_list) * [[]]
                for j in range(len(n_col_list)):
                    item = n_col_list[j][i]
                    temporary_output[j] = item
                simple_list.append(temporary_output)
            return simple_list

        for m in range(self.num_of_ges):
            name = '/gesture%d_exclusive.xlsx' % m
            ws = load_workbook(self.data_folder_path + name).active
            output_filter = ws.max_column * [[]]
            cnt = 0
            for col in ws.iter_cols(min_col=1, max_col=ws.max_column, min_row=2, max_row=ws.max_row, values_only=True):
                output_filter[cnt] = band_filter(col)
                cnt += 1
            output_filter = n_col_list2simple_list(output_filter)
            data_list.append(output_filter)

        for ges in range(self.num_of_ges):
            idx = 0
            while (idx + 2) * 100 <= len(data_list[ges]):
                for j in range(100):
                    x_train.append(np.expand_dims(np.array(
                        data_list[ges][(idx * 100 + j): ((idx + 1) * 100 + j)]).T, axis=0))
                    y_train.append(ges)
                idx += 1

        self.array_x = np.array(x_train)
        self.array_y = np.array(y_train)

    def training(self):
        accuracy = []
        for iter in range(1):
            indices = np.arange(len(self.array_x))
            np.random.shuffle(indices)
            x = self.array_x[indices]
            y = self.array_y[indices]

            x_train = x[0:int(len(self.array_x) * 1)]
            y_train = y[0:int(len(self.array_y) * 1)]
            x_validation = x[int(len(self.array_x) * 0.8):]
            y_validation = y[int(len(self.array_y) * 0.8):]

            dataset_train = TensorDataset(torch.from_numpy(np.array(x_train, dtype=np.float32)),
                                          torch.from_numpy(np.array(y_train, dtype=np.int64)))
            dataset_validation = TensorDataset(torch.from_numpy(np.array(x_validation, dtype=np.float32)),
                                               torch.from_numpy(np.array(y_validation, dtype=np.int64)))
            dataset_test = TensorDataset(torch.from_numpy(np.array(self.array_x, dtype=np.float32)),
                                         torch.from_numpy(np.array(self.array_y, dtype=np.int64)))

            train_loader = torch.utils.data.DataLoader(dataset_train,
                                                       batch_size=self.batch_size, shuffle=True, drop_last=True)
            validation_loader = torch.utils.data.DataLoader(dataset_validation,
                                                            batch_size=self.batch_size, shuffle=True, drop_last=True)
            test_loader = torch.utils.data.DataLoader(dataset_test, batch_size=self.batch_size, shuffle=False)

            dnn = DeepNeuralNet(self.num_of_ges, batch_size=self.batch_size).to(self.device)
            precision = 1e-6
            criterion = nn.CrossEntropyLoss(size_average=False)
            optimizer = optim.Adam(dnn.parameters(), lr=self.learning_rate)
            scheduler = optim.lr_scheduler.ReduceLROnPlateau(optimizer=optimizer, mode='min', factor=0.2, patience=5,
                                                             verbose=True, eps=precision)
            dnn = self.train_model(dnn, criterion, optimizer, scheduler,
                                   data_loaders={"train": train_loader,
                                                 "val": validation_loader},
                                   precision=precision, num_epochs=self.training_iter)
            dnn.eval()

            total = 0
            correct_prediction_test = 0
            for k, data_test in enumerate(test_loader, 0):
                inputs_test, ground_truth_test = data_test
                inputs_test, ground_truth_test = Variable(inputs_test.to(self.device)), Variable(
                    ground_truth_test.to(self.device))
                outputs_test = dnn(inputs_test)
                _, predicted = torch.max(outputs_test.data, 1)
                correct_prediction_test += (predicted.cpu().numpy() == ground_truth_test.data.cpu().numpy()).sum()
                total += ground_truth_test.size(0)
            accuracy.append(correct_prediction_test / total)
            torch.save(dnn, 'dnn.pt')
        print('Overall Accuracy: ', np.mean(accuracy))
        return np.mean(accuracy)

    def train_model(self, dnn, criterion, optimizer, scheduler, data_loaders, precision, num_epochs):

        best_loss = float('inf')
        best_weights = copy.deepcopy(dnn.state_dict())

        for epoch in range(num_epochs):
            print('\nEpoch: ', epoch)
            for phase in ['train', 'val']:
                if phase == 'train':
                    dnn.train(True)
                else:
                    dnn.train(False)

                running_loss = 0.
                running_corrects = 0
                total = 0

                for i, data in enumerate(data_loaders[phase], 0):
                    inputs, labels = data
                    inputs, labels = Variable(inputs.to(self.device)), Variable(labels.to(self.device))

                    optimizer.zero_grad()
                    if phase == 'train':
                        dnn.train()
                        outputs = dnn(inputs)
                        _, predictions = torch.max(outputs.data, 1)
                        loss = criterion(outputs, labels)
                        loss.backward()
                        optimizer.step()
                        loss = loss.item()
                    else:
                        dnn.eval()
                        outputs = dnn(inputs)
                        _, predictions = torch.max(outputs.data, 1)
                        loss = criterion(outputs, labels)
                        loss = loss.item()
                    running_loss += loss
                    running_corrects += torch.sum(predictions == labels.data)
                    total += labels.size(0)

                epoch_loss = running_loss / total
                epoch_acc = running_corrects.item() / total

                print('{} Loss: {} Epoch Accuracy: {}'.format(phase, epoch_loss, epoch_acc))
                if phase == 'val':
                    scheduler.step(epoch_loss)
                    if epoch_loss + precision < best_loss:
                        print("New best validation loss:", epoch_loss)
                        best_loss = epoch_loss
                        best_weights = copy.deepcopy(dnn.state_dict())

        dnn.load_state_dict(copy.deepcopy(best_weights))
        dnn.eval()
        return dnn


if __name__ == '__main__':
    data_path = os.getcwd() + '/train'
    param = {
        'num_of_ges': 8,
        'threshold': 2,
        'window_length': 20,
        'visualization': False,
        'training_iter': 5,
        'learning_rate': 0.002,
        'batch_size': 16
    }

    TrainDNN = TrainDNN(data_path, param)
    TrainDNN.data_preprocessing()
    TrainDNN.training()
