# coding = utf8

import os
import random

import numpy as np
import xgboost as xgb
import matplotlib.pyplot as plt
import sklearn.model_selection

from openpyxl import load_workbook

from online_ges_detection_xgb import band_filter
from online_ges_detection_xgb import segmentation
from online_ges_detection_xgb import get_moving_avg
from online_ges_detection_xgb import rebuild_channel_data


class TrainXGB(object):
    def __init__(self, data_folder_path, param):
        self.data_folder_path = data_folder_path
        self.num_of_ges = param['num_of_ges']
        self.threshold = param['threshold']
        self.window_length = param['window_length']
        self.visualization = param['visualization']
        self.training_iter = param['training_iter']
        self.array_x = []
        self.array_y = []

    def data_preprocessing(self):
        data_list = []
        data_dict = {}
        feature_vec = []

        def n_col_list2simple_list(n_col_list):
            minimum_row = min([len(n_col_list[i]) for i in range(len(n_col_list))])
            simple_list = []
            for i in range(minimum_row):
                temporary_output = len(n_col_list) * [[]]
                for j in range(len(n_col_list)):
                    item = n_col_list[j][i]
                    temporary_output[j] = item
                simple_list.append(temporary_output)
            return simple_list

        def get_feature_mav(data_cut):
            """
            Input:
                data_cut (list): dim = num_of_cut x num_of_sample_per_cut x num_of_channel
            Output:
                mav (array): dim = num_of_cut x num_of_channel
            """
            num_of_cut = len(data_cut)
            num_of_channel = len(data_cut[0][0])
            mav = np.zeros([num_of_cut, num_of_channel])
            for i in range(len(data_cut)):
                mav[i] = np.mean(data_cut[i], axis=0)
            return mav

        for m in range(self.num_of_ges):
            name = '/training%d.xlsx' % m
            ws = load_workbook(self.data_folder_path + name).active
            output_filter = ws.max_column * [[]]
            cnt = 0
            for col in ws.iter_cols(min_col=1, max_col=ws.max_column, min_row=2, max_row=ws.max_row, values_only=True):
                output_filter[cnt] = band_filter(col)
                cnt += 1
            output_filter = n_col_list2simple_list(output_filter)
            data_list.append(output_filter)

        for ges in range(self.num_of_ges):
            data_dict['ges%d' % ges] = rebuild_channel_data(data_list[ges])
            data_dict['ges%d_processed' % ges] = get_moving_avg(data_dict['ges%d' % ges])
            data_dict['ges%d_start' % ges], data_dict['ges%d_end' % ges] = \
                segmentation(data_dict['ges%d_processed' % ges], self.threshold, self.window_length)

            num = len(data_dict['ges%d_start' % ges])
            data_cut = [None] * len(data_dict['ges%d_start' % ges])
            for i in range(len(data_dict['ges%d_start' % ges])):
                data_cut[i] = data_list[ges][data_dict['ges%d_start' % ges][i]:data_dict['ges%d_end' % ges][i]]

            if self.visualization:
                plt.figure(figsize=(50, 3))
                for i in range(1, num + 1):
                    plt.subplot2grid((1, num + 1), (0, i - 1), colspan=1).plot(data_cut[i - 1])
                    plt.title(ges)
                plt.show()

            mav = np.append(get_feature_mav(data_cut), len(data_cut) * [[ges]], axis=1).tolist()
            feature_vec += mav
        self.array_x = np.array(feature_vec)[:, :-1]
        self.array_y = np.array(feature_vec)[:, -1]


    def training(self):
        bst_now = None
        max_accuracy = 0
        accumulated_accuracy = 0
        for numK in range(self.training_iter):
            seed = random.randint(1001, 500900)
            test_size = 0.2

            x_train, x_test, y_train, y_test = sklearn.model_selection.train_test_split(self.array_x, self.array_y,
                                                                                        test_size=test_size,
                                                                                        random_state=seed)

            xg_train = xgb.DMatrix(x_train, label=y_train)
            xg_test = xgb.DMatrix(x_test)
            num_round = 200
            param = {
                'objective': 'multi:softmax',
                'num_class': 8,
                'max_depth': 5,
                'eta': 0.01,
                'nthread': 4,
            }

            bst = xgb.train(param, xg_train, num_round)
            bst.save_model('xgb.model')

            prediction = bst.predict(xg_test)
            cnt1 = 0
            cnt2 = 0
            for i in range(len(x_test)):
                if int(prediction[i]) == int(y_test[i]):
                    cnt1 += 1
                else:
                    cnt2 += 1
            accuracy = (100 * cnt1 / (cnt1 + cnt2))
            if accuracy > max_accuracy:
                max_accuracy = accuracy
                bst_now = bst
            accumulated_accuracy += accuracy
            print(("Accuracy: %.2f %% " % accuracy))
        accumulated_accuracy = accumulated_accuracy / 100.0
        return accumulated_accuracy, bst_now


if __name__ == '__main__':
    data_path = os.getcwd() + '/train'
    param = {
        'num_of_ges': 8,
        'threshold': 2,
        'window_length': 20,
        'visualization': False,
        'training_iter': 100
    }

    TrainXGB = TrainXGB(data_path, param)
    TrainXGB.data_preprocessing()
    TrainXGB.training()
